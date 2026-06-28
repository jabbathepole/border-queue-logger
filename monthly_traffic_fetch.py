"""
Border-Guard monthly traffic BASELINE fetcher (Series A enrichment).

dane.gov.pl dataset 2708, per-crossing family "ruch graniczny ... z podziałem na
odcinki i przejścia (GDDKiA)". Pull-on-demand, NOT a live logger. Discovers the
current per-crossing resource (never hardcodes an id — they change as new files
publish), downloads + ARCHIVES the raw XLSX, maps the multi-row merged header,
extracts the `z Ukrainą` section, and UPSERTs into data/monthly_traffic.db.

Run manually:  python monthly_traffic_fetch.py            # latest available year
               python monthly_traffic_fetch.py --year 2025
               python monthly_traffic_fetch.py --all-years # full backfill

Does NOT touch scraper.py / queue_records or any running logger. Licence CC0 1.0
(Komenda Główna Straży Granicznej via dane.gov.pl 2708). See RECON handoff +
METHODOLOGY.md. The XLSX layout is the fragile part: header positions are
asserted, and ANY mismatch hard-fails (it does not guess columns).
"""
import argparse
import datetime
import io
import logging
import os
import re
import sys
import time

import requests
from openpyxl import load_workbook

from crossings import CANONICAL_NAMES
from monthly_traffic_db import export_csv, init_db, upsert_records

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("monthly_traffic_fetch.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

DATASET_ID = "2708"
RESOURCES_URL = f"https://api.dane.gov.pl/1.4/datasets/{DATASET_ID}/resources?per_page=100"
FILE_URL = "https://api.dane.gov.pl/resources/{rid}/file"
DB_PATH = "data/monthly_traffic.db"
RAW_DIR = "data/raw_traffic"
INCIDENTS = "INCIDENTS.md"

# Selects the per-crossing (GDDKiA) family from the dataset's other resources
# (e.g. the "sam. ciężarowe + razem" bunasta family, which is by-section only).
FAMILY_TITLE_MARKER = "z podziałem na odcinki i przejścia"

USER_AGENT = (
    "border-queue-logger/0.1 "
    "(+https://github.com/jabbathepole/border-queue-logger; "
    "contact d.jablonski97@gmail.com)"
)
HEADERS = {"User-Agent": USER_AGENT}

MAX_ATTEMPTS = 4
BACKOFF_SECONDS = 15
REQUEST_TIMEOUT = 60

ROMAN_MONTHS = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
    "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12,
}

# Vehicle-category header labels expected across the top of every monthly sheet,
# left-to-right. The whole header is asserted (not just the blocks we extract) so
# a layout change anywhere surfaces. label -> our vehicle_type ('' = not mapped).
CATEGORY_LABELS = [
    "RAZEM (łączny ruch graniczny)",
    "pozostałe - inne śr. transportu",
    "motocykle",
    "autobusy",
    "samochody osobowe",
    "samochody ciężarowe",
]
CATEGORY_TO_VEHICLE_TYPE = {
    "RAZEM (łączny ruch graniczny)": "total",
    "autobusy": "bus",
    "samochody osobowe": "car",
    "samochody ciężarowe": "truck",
}
# Populated by THIS PR. The column map is built for every mapped category above
# so cars/buses are a one-line change later; only these are written now.
TARGET_VEHICLE_TYPES = {"total", "truck"}

# Border-section headers in column A (normalised). UA is the only one we ingest.
SECTION_RU = "z federacją rosyjską"
SECTION_BY = "z republiką białoruś"
SECTION_UA = "z ukrainą"
SECTION_HEADERS = {SECTION_RU, SECTION_BY, SECTION_UA}

# Explicit crossing map (recon: dataset 2090 is stale, this file is the authority).
# Keys are normalised on load; the source uses en-dashes + footnote markers, so
# matching is done on _norm_crossing() of both sides.
TRAFFIC_LABEL_TO_CANONICAL = {
    "Budomierz – Hruszew":     "budomierz",
    "Dołhobyczów – Uhrynów":   "dolhobyczow",
    "Dorohusk – Jagodzin":     "dorohusk",
    "Hrebenne – Rawa Ruska":   "hrebenne",
    "Korczowa – Krakowiec":    "korczowa",
    "Krościenko – Smolnica":   "kroscienko",
    "Malhowice – Niżankowice": "malhowice",  # see joint-reporting below
    "Medyka – Szeginie":       "medyka",
    "Zosin – Ustiług":         "zosin",
}


class ResourceDiscoveryError(RuntimeError):
    """No per-crossing resource matched the family/year. The source naming may
    have shifted; surfaces rather than silently using a stale id."""


class LayoutError(RuntimeError):
    """A header label is not where the column map expects it. Hard-fails so a
    workbook layout change is caught instead of mis-reading columns."""


class UnknownTrafficCrossingError(RuntimeError):
    """A `z Ukrainą` row label that resolves to no canonical crossing. Hard-fails
    (same philosophy as the DPSU logger) so a renamed/new crossing surfaces."""


# --------------------------------------------------------------------------- #
# normalisation
# --------------------------------------------------------------------------- #
def _norm(s) -> str:
    """Collapse whitespace + strip; for header-label comparison."""
    return re.sub(r"\s+", " ", str(s or "").replace("\n", " ")).strip()


def _norm_crossing(s) -> str:
    """Defensive crossing-label key: drop trailing footnote markers (`***)`),
    unify dash variants, collapse whitespace, lowercase."""
    t = str(s or "").strip()
    t = re.sub(r"[\s\*]+\)?\s*$", "", t)        # trailing ' ***)' / ' *'
    t = re.sub(r"[‐-―−]", "-", t)  # –, —, − ... -> -
    t = re.sub(r"\s*-\s*", " - ", t)
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


_CANONICAL_BY_NORM = {_norm_crossing(k): v for k, v in TRAFFIC_LABEL_TO_CANONICAL.items()}


def _cell_int(v):
    """Numeric cell -> int; blank/non-numeric -> None."""
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- #
# resource discovery (pure, offline-testable)
# --------------------------------------------------------------------------- #
def _resource_year(title: str) -> int | None:
    years = re.findall(r"20\d\d", title or "")
    return int(years[0]) if years else None


def per_crossing_resources(api_data: list) -> list[dict]:
    """From a dataset /resources `data` list, the per-crossing (GDDKiA) family
    only, each annotated with its parsed data year."""
    out = []
    for r in api_data:
        a = r.get("attributes", {})
        title = a.get("title", "")
        if FAMILY_TITLE_MARKER.lower() not in title.lower():
            continue
        if (a.get("format") or "").lower() not in ("xlsx", "xls"):
            continue
        out.append({
            "id": str(r.get("id")),
            "title": title,
            "year": _resource_year(title),
            "created": a.get("created", ""),
        })
    return out


def available_years(api_data: list) -> list[int]:
    return sorted({r["year"] for r in per_crossing_resources(api_data) if r["year"]})


def select_resource(api_data: list, year: int) -> dict:
    """The latest per-crossing resource for `year` (max created). Hard-fails if
    none — never falls back to a stale id."""
    fam = [r for r in per_crossing_resources(api_data) if r["year"] == year]
    if not fam:
        avail = available_years(api_data)
        raise ResourceDiscoveryError(
            f"No per-crossing (GDDKiA) resource for year {year} in dataset "
            f"{DATASET_ID}. Family marker {FAMILY_TITLE_MARKER!r}; available "
            f"years={avail}. Source naming may have changed — inspect "
            f"{RESOURCES_URL} and record an INCIDENTS.md entry."
        )
    return max(fam, key=lambda r: r["created"])


# --------------------------------------------------------------------------- #
# network
# --------------------------------------------------------------------------- #
def _get(url: str, *, binary: bool):
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            return resp.content if binary else resp.json()
        except Exception as exc:
            if attempt == MAX_ATTEMPTS:
                log.error("GET %s failed after %d attempts: %s", url, MAX_ATTEMPTS, exc)
                raise
            wait = BACKOFF_SECONDS * attempt
            log.warning("GET attempt %d/%d failed (%s) — retry in %ds",
                        attempt, MAX_ATTEMPTS, exc, wait)
            time.sleep(wait)


def fetch_resource_list() -> list:
    return _get(RESOURCES_URL, binary=False)["data"]


def download_and_archive(resource_id: str, fetched_at: str) -> bytes:
    """Download the XLSX and archive it verbatim BEFORE parsing (provenance /
    restatement history)."""
    content = _get(FILE_URL.format(rid=resource_id), binary=True)
    os.makedirs(RAW_DIR, exist_ok=True)
    stamp = fetched_at.replace(":", "").replace("-", "")
    path = os.path.join(RAW_DIR, f"{DATASET_ID}_{resource_id}_{stamp}.xlsx")
    with open(path, "wb") as f:
        f.write(content)
    log.info("Archived raw XLSX -> %s (%d bytes)", path, len(content))
    return content


# --------------------------------------------------------------------------- #
# workbook layout: build column map, then ASSERT it
# --------------------------------------------------------------------------- #
def build_column_map(ws) -> dict:
    """Locate the 3-row header (category / registration / direction) and return
    the column index of every (vehicle_type, registration, direction) cell.
    Raises LayoutError if any expected label is absent or misplaced.

    Each category is a 5-column block: [RAZEM | obce(do RP, z RP) | polskie(do
    RP, z RP)]. We assert the whole header (all 6 categories), not just the
    blocks we extract, so any layout drift is caught.
    """
    # 1. find the category row (the one carrying 'samochody ciężarowe')
    cat_row = None
    for r in range(1, 9):
        labels = {_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if "samochody ciężarowe" in labels:
            cat_row = r
            break
    if cat_row is None:
        raise LayoutError(
            "Category header row (containing 'samochody ciężarowe') not found in "
            "the first 8 rows — workbook layout changed."
        )
    reg_row, dir_row = cat_row + 1, cat_row + 2

    # 2. locate each expected category's start column
    starts: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        lbl = _norm(ws.cell(cat_row, c).value)
        if lbl in CATEGORY_LABELS:
            starts[lbl] = c
    missing = [lbl for lbl in CATEGORY_LABELS if lbl not in starts]
    if missing:
        raise LayoutError(
            f"Expected category headers not found on row {cat_row}: {missing}. "
            f"Found: {sorted(starts)}."
        )

    # 3. assert each block's registration + direction sub-headers, build the map
    vehicles: dict[str, dict] = {}
    for lbl, s in starts.items():
        def need(row, col, expect, *, prefix=False):
            got = _norm(ws.cell(row, col).value).lower()
            exp = expect.lower()
            ok = got.startswith(exp) if prefix else got == exp
            if not ok:
                raise LayoutError(
                    f"Header mismatch in block {lbl!r} at row {row} col {col}: "
                    f"expected {expect!r}, got {ws.cell(row, col).value!r}."
                )
        need(reg_row, s, "RAZEM")
        need(reg_row, s + 1, "obce")
        need(reg_row, s + 3, "polskie")
        need(dir_row, s + 1, "do RP", prefix=True)
        need(dir_row, s + 2, "z RP", prefix=True)
        need(dir_row, s + 3, "do RP", prefix=True)
        need(dir_row, s + 4, "z RP", prefix=True)

        vt = CATEGORY_TO_VEHICLE_TYPE.get(lbl)
        if vt:
            # REGISTRATION EXTENSION POINT. "total" is the category's single
            # all-registrations-all-directions cell. "foreign"/"polish" are the
            # per-direction columns. registration='all' = foreign+polish (below).
            # MUTUAL EXCLUSIVITY: 'all' already == foreign+polish, so a future
            # split that emits 'foreign'/'polish' rows must NEVER be summed
            # together with 'all' — doing so doubles every total.
            vehicles[vt] = {
                "total": s,
                "foreign": {"do_RP": s + 1, "z_RP": s + 2},
                "polish":  {"do_RP": s + 3, "z_RP": s + 4},
            }

    return {
        "cat_row": cat_row, "reg_row": reg_row, "dir_row": dir_row,
        "data_start_row": dir_row + 1, "vehicles": vehicles,
    }


# --------------------------------------------------------------------------- #
# extraction
# --------------------------------------------------------------------------- #
def _sheet_year(ws, fallback: int) -> int:
    m = re.search(r"20\d\d", _norm(ws.cell(1, 1).value))
    return int(m.group()) if m else fallback


def _extract_crossing_values(ws, row: int, cm: dict) -> dict:
    """For one data row, the registration='all' (foreign+polish) count for each
    target vehicle_type × direction. Also returns a soft integrity check against
    the category RAZEM total column."""
    values: dict[tuple, int | None] = {}
    integrity_ok = True
    for vt, cols in cm["vehicles"].items():
        if vt not in TARGET_VEHICLE_TYPES:
            continue
        dir_all: dict[str, int | None] = {}
        for direction in ("z_RP", "do_RP"):
            f = _cell_int(ws.cell(row, cols["foreign"][direction]).value)
            p = _cell_int(ws.cell(row, cols["polish"][direction]).value)
            # registration='all' for this direction = foreign(obce) + Polish(polskie).
            # If a future extension also stores 'foreign'/'polish' rows, treat them
            # as a SEPARATE view of this same number — never aggregate 'all' with
            # 'foreign'/'polish' (it would double-count). See METHODOLOGY.md.
            dir_all[direction] = None if (f is None and p is None) else (f or 0) + (p or 0)
            values[(vt, direction)] = dir_all[direction]
        # integrity: RAZEM (all dir, all reg) should equal z_RP_all + do_RP_all
        total = _cell_int(ws.cell(row, cols["total"]).value)
        recomputed = (dir_all["z_RP"] or 0) + (dir_all["do_RP"] or 0)
        if total is not None and total != recomputed:
            integrity_ok = False
    values["_integrity_ok"] = integrity_ok
    return values


def extract_sheet(ws, resource_id: str, fetched_at: str, fallback_year: int) -> list[dict]:
    """Extract the `z Ukrainą` section of one MONTHLY sheet into records.
    Returns [] for an entirely-zero sheet (a not-yet-published month — so a later
    pull can fill it; genuine zeros within a published month are kept as 0)."""
    cm = build_column_map(ws)
    year = _sheet_year(ws, fallback_year)
    month_num = ROMAN_MONTHS[ws.title]
    month = f"{year}-{month_num:02d}"

    # find the UA section start
    ua_row = None
    for r in range(cm["data_start_row"], ws.max_row + 1):
        if _norm(ws.cell(r, 1).value).lower() == SECTION_UA:
            ua_row = r
            break
    if ua_row is None:
        raise LayoutError(f"'z Ukrainą' section header not found on sheet {ws.title!r}.")

    # collect raw per-crossing values until the section ends
    raw: dict[str, dict] = {}     # crossing_id -> {(vt,dir): val, "_label":, "_integrity_ok":}
    for r in range(ua_row + 1, ws.max_row + 1):
        label_raw = ws.cell(r, 1).value
        norm = _norm_crossing(label_raw)
        if not norm:
            break                                   # blank row -> end of section
        if norm.startswith("*") or "odprawa poza" in norm or "suma razem" in norm \
                or norm.startswith("spr"):
            break                                   # footnote / summary row -> end
        if norm in SECTION_HEADERS:
            break                                   # next border section -> end
        cid = _CANONICAL_BY_NORM.get(norm)
        if cid is None:
            raise UnknownTrafficCrossingError(
                f"Unknown 'z Ukrainą' crossing label {str(label_raw)!r} (norm "
                f"{norm!r}) on sheet {ws.title!r}: not in TRAFFIC_LABEL_TO_CANONICAL. "
                f"A renamed/new crossing — add it and record an INCIDENTS.md entry."
            )
        vals = _extract_crossing_values(ws, r, cm)
        vals["_label"] = _norm(label_raw)
        raw[cid] = vals

    if not vals_integrity(raw):
        log.warning("[%s] RAZEM integrity check failed for >=1 crossing — column "
                    "map may have drifted; inspect headers.", month)

    # not-yet-published month: every count zero/None -> skip the whole sheet
    if not _has_positive(raw):
        log.info("[%s] all-zero sheet — treating as not-yet-published, skipped.", month)
        return []

    raw = apply_joint_reporting(raw, month)

    records: list[dict] = []
    for cid, vals in raw.items():
        joint = vals.get("_joint_with")
        for vt in sorted(TARGET_VEHICLE_TYPES):
            if vt not in {k[0] for k in vals if isinstance(k, tuple)}:
                continue
            for direction in ("z_RP", "do_RP"):
                count = None if joint else vals.get((vt, direction))
                records.append({
                    "crossing_id": cid,
                    "crossing_label": vals["_label"],
                    "month": month,
                    "vehicle_type": vt,
                    "direction": direction,
                    # 'all' == foreign+polish; mutually exclusive with any future
                    # 'foreign'/'polish' rows — never sum across registration.
                    "registration": "all",
                    "count": count,
                    "joint_reported_with": joint,
                    "source_resource_id": resource_id,
                    "source_dataset": DATASET_ID,
                    "fetched_at": fetched_at,
                })
    log.info("[%s] %d crossings -> %d rows", month, len(raw), len(records))
    return records


def vals_integrity(raw: dict) -> bool:
    return all(v.get("_integrity_ok", True) for v in raw.values())


def _has_positive(raw: dict) -> bool:
    for vals in raw.values():
        for k, v in vals.items():
            if isinstance(k, tuple) and isinstance(v, int) and v > 0:
                return True
    return False


def apply_joint_reporting(raw: dict, month: str) -> dict:
    """Decide whether Małhowice is independent of Medyka for this month.

    REALITY (verified against the live file 2026-06-28): in the MONTHLY sheets we
    ingest, Małhowice–Niżankowice and Medyka–Szeginie are genuinely SEPARATE —
    Małhowice is a small passenger crossing (real car traffic, typically zero
    trucks), Medyka is the large one. So the normal, correct behaviour is to store
    both INDEPENDENTLY. (The identical values recon saw were in the *annual* Razem
    aggregate sheet, which duplicates Medyka's total into the Małhowice row — and
    we never ingest that sheet.)

    GUARD (the brief's double-count concern): if the two rows are nonetheless
    IDENTICAL across every captured field — i.e. the annual-sheet duplication has
    leaked into a monthly sheet — then Małhowice is NOT real independent data:
    collapse it (store the figure under medyka only, malhowice count=NULL,
    joint_reported_with='medyka') so a sum never double-counts, and record the
    anomaly once. This is the "verify, don't assume" check the brief asked for —
    just with default and exception swapped to match the real data.
    """
    mal, med = raw.get("malhowice"), raw.get("medyka")
    if not mal or not med:
        return raw

    keys = [k for k in mal if isinstance(k, tuple)]
    identical = bool(keys) and all(mal.get(k) == med.get(k) for k in keys)
    if identical:
        mal["_joint_with"] = "medyka"
        log.warning("[%s] malhowice == medyka (identical values) — upstream "
                    "duplication; treating as joint-reported, malhowice count=NULL.", month)
        _record_joint_incident(month)
    else:
        mal["_joint_with"] = None  # normal: genuinely separate crossings
    return raw


def _record_joint_incident(month: str) -> None:
    """Idempotently note an identical-values (joint-reported) month in
    INCIDENTS.md. Best-effort; never fatal; written once per month."""
    marker = f"monthly_traffic: Małhowice==Medyka identical ({month})"
    try:
        if os.path.exists(INCIDENTS):
            if marker in open(INCIDENTS, encoding="utf-8").read():
                return
        entry = (
            f"\n\n## {marker}\n\n"
            f"Detected {datetime.datetime.now(datetime.timezone.utc):%Y-%m-%dT%H:%M:%SZ}. "
            f"Dataset 2708 reported Małhowice–Niżankowice and Medyka–Szeginie with "
            f"IDENTICAL values for {month} — the annual-aggregate duplication has "
            f"leaked into a monthly sheet. Małhowice is therefore stored as "
            f"joint-reported (count=NULL, joint_reported_with='medyka') to avoid a "
            f"double-count. If Małhowice should have its own figures this month, "
            f"the source needs correcting.\n"
        )
        with open(INCIDENTS, "a", encoding="utf-8") as f:
            f.write(entry)
    except OSError as exc:
        log.warning("Could not append INCIDENTS.md entry: %s", exc)


def parse_workbook(content: bytes, resource_id: str, fetched_at: str,
                   fallback_year: int) -> list[dict]:
    """Parse all MONTHLY sheets (I–XII). Quarter/half/Razem aggregate sheets are
    skipped (exact roman-numeral sheet names only) to avoid double-counting."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records: list[dict] = []
    monthly = [s for s in wb.sheetnames if s in ROMAN_MONTHS]
    log.info("Monthly sheets: %s (skipping %d aggregate sheets)",
             monthly, len(wb.sheetnames) - len(monthly))
    for name in monthly:
        records.extend(extract_sheet(wb[name], resource_id, fetched_at, fallback_year))
    wb.close()
    return records


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #
def run(years: list[int] | None) -> int:
    fetched_at = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    api_data = fetch_resource_list()

    if years is None:
        avail = available_years(api_data)
        if not avail:
            raise ResourceDiscoveryError("No per-crossing resources discovered at all.")
        years = [max(avail)]
    log.info("Target years: %s", years)

    all_records: list[dict] = []
    for year in years:
        res = select_resource(api_data, year)
        log.info("Year %d -> resource %s (%s)", year, res["id"], res["title"])
        content = download_and_archive(res["id"], fetched_at)
        all_records.extend(parse_workbook(content, res["id"], fetched_at, year))

    if not all_records:
        log.warning("No records extracted (all targeted months unpublished?).")
        return 0

    init_db(DB_PATH)
    stats = upsert_records(DB_PATH, all_records)
    log.info("UPSERT: %d inserted, %d updated (of %d rows)",
             stats["inserted"], stats["updated"], len(all_records))
    csv_path = export_csv(DB_PATH)
    if csv_path:
        log.info("CSV snapshot -> %s", csv_path)
    return len(all_records)


def main() -> None:
    ap = argparse.ArgumentParser(description="Fetch dataset 2708 monthly traffic baseline.")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--year", type=int, help="Single data year to pull.")
    g.add_argument("--all-years", action="store_true", help="Pull every available year.")
    args = ap.parse_args()

    log.info("=== monthly_traffic fetch starting ===")
    try:
        if args.all_years:
            years = available_years(fetch_resource_list())
        elif args.year:
            years = [args.year]
        else:
            years = None  # latest available
        run(years)
    except (ResourceDiscoveryError, LayoutError, UnknownTrafficCrossingError) as exc:
        log.error("Aborted — %s", exc)
        sys.exit(1)
    except Exception:
        log.exception("Aborted — unexpected error")
        sys.exit(1)
    log.info("=== Run complete ===")


if __name__ == "__main__":
    main()
