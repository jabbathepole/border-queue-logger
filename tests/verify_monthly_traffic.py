"""Offline checks for the monthly-traffic baseline (dataset 2708). No network:
builds in-memory XLSX fixtures matching the real merged-header layout.

  1. Resource discovery: per-crossing family filter, latest-per-year selection,
     and a hard-fail (ResourceDiscoveryError) when no resource matches the year.
  2. _norm_crossing() handles en-dashes + footnote markers; all 9 UA labels map.
  3. build_column_map() locates the truck/total blocks; a mutated header raises
     LayoutError (no silent column mis-read).
  4. extract_sheet() end-to-end: 9 crossings -> rows, both directions, RAZEM
     integrity, and a hard-fail on an unknown crossing label.
  5. Joint reporting: identical malhowice/medyka -> malhowice count=NULL,
     joint_reported_with='medyka'; divergence -> independent + INCIDENTS entry.
  6. Not-yet-published: an all-zero sheet is skipped.
  7. UPSERT: re-pull is idempotent; a backfilled month overwrites a prior zero.

Pure offline; safe under pytest or `python tests/verify_monthly_traffic.py`.
"""
import os
import sqlite3
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import monthly_traffic_fetch as mtf
from crossings import CANONICAL_NAMES
from monthly_traffic_db import init_db, upsert_records
from monthly_traffic_fetch import (
    LayoutError,
    ResourceDiscoveryError,
    UnknownTrafficCrossingError,
    _norm_crossing,
    available_years,
    build_column_map,
    extract_sheet,
    select_resource,
)

failures: list[str] = []

# Redirect INCIDENTS.md writes to a throwaway file for the whole run so no test
# fixture (several use identical malhowice/medyka) ever touches the real file.
# Section 5 overrides this locally where it asserts on the written content.
_GLOBAL_INC = tempfile.NamedTemporaryFile("w", suffix=".md", delete=False, encoding="utf-8")
_GLOBAL_INC.close()
mtf.INCIDENTS = _GLOBAL_INC.name


def check(label: str, cond: bool, detail: str = "") -> None:
    print(f"{'ok  ' if cond else 'FAIL'} {label}{(' — ' + detail) if detail and not cond else ''}")
    if not cond:
        failures.append(label)


# --------------------------------------------------------------------------- #
# fixture builder: replicate the real 3-row merged header + a UA section
# --------------------------------------------------------------------------- #
# category start columns (1-indexed): B,G,L,Q,V,AA — 5-col blocks
CAT_STARTS = {
    "RAZEM (łączny ruch graniczny)": 2,
    "pozostałe - inne śr. transportu": 7,
    "motocykle": 12,
    "autobusy": 17,
    "samochody osobowe": 22,
    "samochody ciężarowe": 27,
}


def _write_header(ws):
    ws.cell(1, 1, "STYCZEŃ 2026 r.")
    for lbl, s in CAT_STARTS.items():
        ws.cell(3, s, lbl)
        ws.cell(4, s, "RAZEM")
        ws.cell(4, s + 1, "obce")
        ws.cell(4, s + 3, "polskie")
        ws.cell(5, s + 1, "do RP  ")
        ws.cell(5, s + 2, "z RP  ")
        ws.cell(5, s + 3, "do RP  ")
        ws.cell(5, s + 4, "z RP  ")


def _write_block(ws, row, start, odo, oz, pdo, pz):
    """obce do/z, polskie do/z; RAZEM = sum (matches the real file)."""
    ws.cell(row, start, odo + oz + pdo + pz)
    ws.cell(row, start + 1, odo)
    ws.cell(row, start + 2, oz)
    ws.cell(row, start + 3, pdo)
    ws.cell(row, start + 4, pz)


def _crossing_row(ws, row, label, truck, total=None):
    """truck/total = (obce_do, obce_z, polskie_do, polskie_z). total defaults to
    2x truck so total != truck (distinct numbers)."""
    ws.cell(row, 1, label)
    _write_block(ws, row, CAT_STARTS["samochody ciężarowe"], *truck)
    t = total or tuple(2 * x for x in truck)
    _write_block(ws, row, CAT_STARTS["RAZEM (łączny ruch graniczny)"], *t)


UA_LABELS = {
    "Budomierz – Hruszew": "budomierz",
    "Dołhobyczów – Uhrynów": "dolhobyczow",
    "Dorohusk – Jagodzin": "dorohusk",
    "Hrebenne – Rawa Ruska": "hrebenne",
    "Korczowa – Krakowiec": "korczowa",
    "Krościenko – Smolnica": "kroscienko",
    "Malhowice – Niżankowice ***)": "malhowice",
    "Medyka – Szeginie": "medyka",
    "Zosin – Ustiług": "zosin",
}


def build_ua_sheet(*, malhowice_joint=True, all_zero=False, title="I") -> "Worksheet":
    wb = Workbook()
    ws = wb.active
    ws.title = title
    _write_header(ws)
    r = 6
    ws.cell(r, 1, "z Ukrainą"); r += 1
    medyka_truck = (100, 200, 30, 40)
    for i, label in enumerate(UA_LABELS):
        if all_zero:
            truck = (0, 0, 0, 0)
        elif label.startswith("Medyka"):
            truck = medyka_truck
        elif label.startswith("Malhowice"):
            truck = medyka_truck if malhowice_joint else (1, 2, 3, 4)
        else:
            truck = (10 + i, 20 + i, 5 + i, 8 + i)
        _crossing_row(ws, r, label, truck)
        r += 1
    ws.cell(r, 1, "*) odprawa poza czynnym przejściem granicznym")  # footnote -> stop
    return ws


# --------------------------------------------------------------------------- #
# 1. resource discovery
# --------------------------------------------------------------------------- #
def _res(rid, title, created, fmt="xlsx"):
    return {"id": rid, "attributes": {"title": title, "created": created, "format": fmt}}


T = "Dane statystyczne - ruch graniczny środków transportu drogowego w dpg z podziałem na odcinki i przejścia"
API = [
    _res(1687171, f"{T} - styczeń-marzec 2026 (GDDKiA)", "2026-04-22T13:17:35Z"),
    _res(1908844, f"{T} - styczeń-kwiecień 2026 (GDDKiA)", "2026-05-21T08:55:13Z"),
    _res(1114803, f"{T} - styczeń-grudzień 2025 r. (GDDKiA)", "2026-01-22T14:03:20Z"),
    _res(99999, "Dane statystyczne - sam. ciężarowe + razem ... 2026 (bunasta)", "2026-05-01T00:00:00Z"),
]
check("available_years finds 2025+2026", available_years(API) == [2025, 2026],
      str(available_years(API)))
check("select latest 2026 = 1908844", select_resource(API, 2026)["id"] == "1908844",
      select_resource(API, 2026)["id"])
check("select 2025 = 1114803", select_resource(API, 2025)["id"] == "1114803")
try:
    select_resource(API, 2099)
    check("no-match year hard-fails", False, "no exception")
except ResourceDiscoveryError:
    check("no-match year hard-fails", True)

# --------------------------------------------------------------------------- #
# 2. normalisation + mapping coverage
# --------------------------------------------------------------------------- #
check("_norm_crossing strips footnote+dash",
      _norm_crossing("Malhowice – Niżankowice ***)") == "malhowice - niżankowice",
      _norm_crossing("Malhowice – Niżankowice ***)"))
mapped = {mtf._CANONICAL_BY_NORM.get(_norm_crossing(k)) for k in UA_LABELS}
check("all 9 UA labels map to canonical", mapped == set(CANONICAL_NAMES),
      f"got {sorted(x for x in mapped if x)}")

# --------------------------------------------------------------------------- #
# 3. column map + layout assertion
# --------------------------------------------------------------------------- #
ws = build_ua_sheet()
cm = build_column_map(ws)
check("truck block columns located",
      cm["vehicles"]["truck"] == {"total": 27, "foreign": {"do_RP": 28, "z_RP": 29},
                                  "polish": {"do_RP": 30, "z_RP": 31}}, str(cm["vehicles"]["truck"]))
check("total block columns located",
      cm["vehicles"]["total"]["foreign"]["z_RP"] == 4, str(cm["vehicles"]["total"]))

ws_bad = build_ua_sheet()
ws_bad.cell(3, 27, "samochody ciezarowe ZMIANA")  # mutate the ciężarowe header
try:
    build_column_map(ws_bad)
    check("mutated header hard-fails", False, "no exception")
except LayoutError:
    check("mutated header hard-fails", True)

# --------------------------------------------------------------------------- #
# 4. extract_sheet end-to-end
# --------------------------------------------------------------------------- #
recs = extract_sheet(build_ua_sheet(), "1908844", "2026-06-28T00:00:00Z", 2026)
by = {}
for rec in recs:
    by.setdefault(rec["crossing_id"], {})[(rec["vehicle_type"], rec["direction"])] = rec
check("9 crossings x 2 vt x 2 dir = 36 rows", len(recs) == 36, str(len(recs)))
check("month parsed YYYY-MM", recs[0]["month"] == "2026-01", recs[0]["month"])
check("registration all", all(r["registration"] == "all" for r in recs))
check("source provenance set",
      recs[0]["source_resource_id"] == "1908844" and recs[0]["source_dataset"] == "2708")
# dorohusk is i=2 in UA_LABELS order -> truck=(12,22,7,10): z_RP_all=22+10=32, do_RP_all=12+7=19
d = by["dorohusk"]
check("dorohusk truck z_RP = 32 (obce_z+polskie_z)", d[("truck", "z_RP")]["count"] == 32,
      str(d[("truck", "z_RP")]["count"]))
check("dorohusk truck do_RP = 19", d[("truck", "do_RP")]["count"] == 19,
      str(d[("truck", "do_RP")]["count"]))

# unknown crossing hard-fails
ws_unk = build_ua_sheet()
ws_unk.cell(8, 1, "Nowe – Przejście")  # row 8 = a crossing row, unmapped
try:
    extract_sheet(ws_unk, "x", "t", 2026)
    check("unknown crossing hard-fails", False, "no exception")
except UnknownTrafficCrossingError:
    check("unknown crossing hard-fails", True)

# --------------------------------------------------------------------------- #
# 5. Małhowice/Medyka handling (default = separate; guard = collapse if identical)
# --------------------------------------------------------------------------- #
# NORMAL (build default builds them identical via malhowice_joint=True ONLY to
# exercise the guard). First: the realistic SEPARATE case = no incident, both kept.
tmp_sep = tempfile.NamedTemporaryFile("w", suffix=".md", delete=False, encoding="utf-8")
tmp_sep.close()
orig_inc = mtf.INCIDENTS
mtf.INCIDENTS = tmp_sep.name
try:
    srecs = extract_sheet(build_ua_sheet(malhowice_joint=False), "x", "t", 2026)
finally:
    mtf.INCIDENTS = orig_inc
sby = {(r["crossing_id"], r["vehicle_type"], r["direction"]): r for r in srecs}
smal = sby[("malhowice", "truck", "z_RP")]
check("separate: malhowice stored independently (count not NULL)", smal["count"] == 6,
      str(smal["count"]))  # (1,2,3,4): z_RP_all = 2+4 = 6
check("separate: joint_reported_with is None", smal["joint_reported_with"] is None)
sep_inc = Path(tmp_sep.name).read_text(encoding="utf-8")
os.unlink(tmp_sep.name)
check("separate: NO INCIDENTS entry written", sep_inc.strip() == "", repr(sep_inc[:80]))

# GUARD: identical values -> collapse malhowice to NULL+joint + record once
check("identical-guard: malhowice count=NULL", by["malhowice"][("truck", "z_RP")]["count"] is None,
      str(by["malhowice"][("truck", "z_RP")]["count"]))
check("identical-guard: malhowice joint_reported_with=medyka",
      by["malhowice"][("truck", "z_RP")]["joint_reported_with"] == "medyka")
check("identical-guard: medyka keeps its value (240)",
      by["medyka"][("truck", "z_RP")]["count"] == 240,
      str(by["medyka"][("truck", "z_RP")]["count"]))  # 200+40

tmp_inc = tempfile.NamedTemporaryFile("w", suffix=".md", delete=False, encoding="utf-8")
tmp_inc.close()
mtf.INCIDENTS = tmp_inc.name
try:
    extract_sheet(build_ua_sheet(malhowice_joint=True), "x", "t", 2026)   # writes once
    extract_sheet(build_ua_sheet(malhowice_joint=True), "x", "t", 2026)   # idempotent
finally:
    mtf.INCIDENTS = orig_inc
inc_text = Path(tmp_inc.name).read_text(encoding="utf-8")
os.unlink(tmp_inc.name)
check("identical-guard: INCIDENTS entry written", "identical" in inc_text.lower(), inc_text[:80])
check("identical-guard: INCIDENTS write is idempotent (one entry)",
      inc_text.count("Małhowice==Medyka identical (2026-01)") == 1,
      str(inc_text.count("Małhowice==Medyka identical (2026-01)")))

# --------------------------------------------------------------------------- #
# 6. not-yet-published all-zero sheet skipped
# --------------------------------------------------------------------------- #
check("all-zero sheet -> [] (unpublished)",
      extract_sheet(build_ua_sheet(all_zero=True), "x", "t", 2026) == [])

# --------------------------------------------------------------------------- #
# 7. UPSERT idempotency + backfill overwrite
# --------------------------------------------------------------------------- #
db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
db.close()
try:
    init_db(db.name)
    base = extract_sheet(build_ua_sheet(), "v1", "2026-06-01T00:00:00Z", 2026)
    s1 = upsert_records(db.name, base)
    s2 = upsert_records(db.name, base)  # re-pull identical vintage
    with sqlite3.connect(db.name) as c:
        n = c.execute("SELECT COUNT(*) FROM monthly_traffic").fetchone()[0]
    check("first upsert inserts 36", s1["inserted"] == 36, str(s1))
    check("re-upsert inserts 0 (idempotent)", s2["inserted"] == 0 and s2["updated"] == 36, str(s2))
    check("no duplicate rows after re-pull", n == 36, str(n))

    # backfill: same keys, new value + new vintage -> overwrite
    for r in base:
        if r["crossing_id"] == "dorohusk" and r["vehicle_type"] == "truck" and r["direction"] == "z_RP":
            r["count"] = 9999
            r["source_resource_id"] = "v2"
    upsert_records(db.name, base)
    with sqlite3.connect(db.name) as c:
        row = c.execute(
            "SELECT count, source_resource_id FROM monthly_traffic WHERE crossing_id='dorohusk' "
            "AND vehicle_type='truck' AND direction='z_RP'").fetchone()
    check("backfill overwrote value + vintage", row == (9999, "v2"), str(row))
finally:
    try:
        os.unlink(db.name)
    except PermissionError:
        pass  # Windows may hold the handle briefly; temp file, harmless

try:
    os.unlink(_GLOBAL_INC.name)
except OSError:
    pass

print()
print("RESULT:", "ALL CHECKS PASS" if not failures else f"{len(failures)} FAILURES: {failures}")
if failures:
    sys.exit(1)
